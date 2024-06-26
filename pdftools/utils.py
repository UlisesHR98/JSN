from enum import Enum
import re

import PyPDF2


class AliasesExtractor:
    def __init__(self, pdf_name, ptype):
        """
        Initialize the PDF parser
        :param pdf_name: name of the pdf to parse
        :param ptype: type of parser. CJ for charles jones, JSN or SC for state capital. JSN = CJ
        """
        self.pdf_file = pdf_name
        self.parser_type = ptype

    def parse_pdf_aliases(self):
        """
        Parse aliases from the PDF set at class initialization
        :return: list of aliases scraped from the pdf
        :rtype: list[str]
        """
        import PyPDF2

        pdf_page_offset = 2
        aliases_list = []
        with open(self.pdf_file, mode='rb') as f:
            reader = PyPDF2.PdfFileReader(f)
            for i, page in enumerate(reader.pages):
                page_text = page.extractText()
                new_text = self._split_and_condense_page_text(page_text)
                report_searched_names = self._extaract_names_searched_in_report(new_text)
                # if report_searched_names:
                #     print("Searched names located\n{}".format(report_searched_names))
                if "ADDED TO OUR INDEX." in page_text:
                    print(('Found addition to index on page {}'.format(i+1)))
                    extracted_aliases = self._extract_aliases_for_judgment(new_text)
                    # aliases_list.extend(extracted_aliases)
        print('Parsing complete')
        # print('Aliases found: {}'.format(aliases_list))
        return list(set(aliases_list))

    def _extaract_names_searched_in_report(self, new_text):
        """
        Pull out the names searched in the original report based on the provided text
        :param new_text: the text to search for search names
        :return: a list of parsed names
        """
        report_searched_names = []
        for line in new_text:
            if self._dates_found_in_line(line):
                # print ("Found search name line:\n{}     ".format(line))
                report_searched_names.append(line)
        return report_searched_names

    def _dates_found_in_line(self, line_text):
        """
        Determine if a date is detected in the provided line of text
        :param line_text: the line of text to parse. extracted from the pdf line
        :return: the name extracted from the line containing a date range based on the search that was conducted in the original pdf
        """
        import re

        extracted_name = None
        regex = '\d{2}\-\d{2}\-\d{4}\s+\d{2}\-\d{2}\-\d{4}'
        m = re.search(regex, line_text)
        if m and len(m.group(0)) > 0:
            extracted_name = self._extract_name_from_name_search_line(line_text)
        return extracted_name

    def _extract_name_from_name_search_line(self, line_text):
        """
        Extract a name from the line of text provided
        :param line_text: the line of text to parse
        :return: the extracted name
        """
        import re
        extracted_name = None
        regex = '^\D+\s'
        m = re.search(regex, line_text)
        if m and len(m.group(0)) > 0:
            extracted_name = m.group(0).strip()
            print(('Search name extracted: {}'.format(extracted_name)))
        return extracted_name


    def _extract_aliases(self, text_lines):
        """
        Extract aliases found in text_lines corresponding to text from a single page
        :param text_lines: the lines of text to parse
        :return: a list of extracted aliases
        """
        aliases = []
        for i, line in enumerate(text_lines):
            if "A/K/A" in line:
                aliases.append(text_lines[i+1])
        return aliases

    def _extract_aliases_for_judgment(self, text_lines):
        """
        Extract aliases and associate with a judgment for matches found in text_lines corresponding to text from a pdf page
        :param text_lines: the lines of text to parse
        :return: a list of extracted aliases
        """
        aliases = []
        for i, line in enumerate(text_lines):
            if "A/K/A" in line:
                # aliases.append(text_lines[i+1])
                alias = text_lines[i+1]
                judgment = self._extract_judgment_for_alias(i, text_lines)
        return aliases

    def _extract_judgment_for_alias(self, line_num, text_lines):
        """
        Extract judgment number from a set of lines. this function parses from the bottom upwards
        :param line_num: the last line number to parse out of the lines of text provided
        :param text_lines: a lis tof text lines to parse
        :return: the matched judgment line text
        """
        judgment_token = 'JUDGMENT NUMBER:'
        judgment_line_text = ''
        for idx in range(line_num, 0, -1):
            if judgment_token in text_lines[idx] and judgment_line_text == '':
                judgment_line_text = text_lines[idx]
        return judgment_line_text

    def _split_and_condense_page_text(self, page_text):
        """
        Clean up text extracted from a pdf by removing empty and blank lines
        :param page_text: the page text to clean up
        :return: the cleaned text
        """
        split_text = page_text.split('\n')
        new_text = []
        for line in split_text:
            if line != "\n" and line != '':
                new_text.append(line.strip())
        return new_text


class JudgmentsFileHash:

    @staticmethod
    def generate_file_hash(filepath):
        """
        Generate a hash (SHA1) for a file
        :param filepath: the file of the path to generate the hash for
        :return: the computed SHA1 hash
        """
        import sys
        import hashlib

        BUF_SIZE = 65536  # 64kb chunks
        sha1 = hashlib.sha1()

        with open(filepath, 'rb') as f:
            while True:
                data = f.read(BUF_SIZE)
                if not data:
                    break
                sha1.update(data)
        return sha1.hexdigest()

    @staticmethod
    def file_already_processed(sha1_hash, processing_type=None):
        """
        Check database for presence of sha1_hash. If found, return true.
        :param sha1_hash: sha1 hash of file
        :return: boolean True if found and already processed
        """
        from pdftools.models import SCDobDocument
        if processing_type == JudgmentsFileHashType.SCDOB:
            d = SCDobDocument.objects.filter(file_hash=sha1_hash)
            if len(d) > 0:
                return True
        return False

class JudgmentsExtractor:

    def __init__(self, pdf_name, ptype):
        """
        Initialize the PDF parser
        :param pdf_name: name of the pdf to parse
        :param ptype: type of parser. CJ for charles jones, JSN or SC for state capital. JSN = CJ
        """
        self.pdf_file = pdf_name
        self.parser_type = ptype

    def parse_pdf_judgments(self):
        """
        Generate a dict of judgments and the pages each judgment was found on
        :return: a dict of judgment numbers and their corresponding page number
        """
        import PyPDF2

        pdf_page_offset = 2
        final_judgment_dict = {}
        with open(self.pdf_file, mode='rb') as f:
            reader = PyPDF2.PdfFileReader(f)
            for i, page in enumerate(reader.pages):
                page_text = page.extractText()
                judgment_page_dict = self._find_judgments_on_page(page_text, i+pdf_page_offset)
                final_judgment_dict.update(judgment_page_dict)
        print('Parsing complete')
        return final_judgment_dict


    def _find_judgments_on_page(self, page_text, pdf_page):
        """
        Parse text for judgments and return matches as dictionary
        :param page_text: extracted text str from pdf
        :param pdf_page: page number of the page being parsed
        :return: dict of matches with judgment numbers as keys
        """
        import re
        d = {}
        regex = None
        judgment_type = self.parser_type
        if judgment_type == "CJ" or judgment_type == "JSN":
            regex = "(\w+\-\w{6}\-\w{4})"
        elif judgment_type == "SC":
            regex = "(.{2}\w{6}\-\d{2})"
        if regex:
            matches = re.findall(regex, page_text)
            for m in matches:
                d[m] = pdf_page
        return d

    @staticmethod
    def _add_extra_judgments_to_workbook(wb, cj_judgments, judgment_idx, cj_idx):
        """

        :param wb: the workbook to update
        :param cj_judgments: a list of CJ judgments to add to the workbook
        :param judgment_idx: the column denoting the JSN judgments
        :param cj_idx: the column denoting where the CJ judgments should be added
        :return:
        """
        jsn_judgments = {}
        ws = wb.worksheets[0]
        for row_num, ws_row in enumerate(ws.rows):  # type: worksheet
            judgment_num = ws_row[judgment_idx].value
            jsn_judgments[judgment_num] = row_num

        for judg in list(cj_judgments.keys()):
            if not judg is None and not jsn_judgments.get(judg):
                # append row to ws
                # csv_header = ['Document Number', 'Matched Names', 'DOB', 'Address', 'Filing Location', 'Case Title',
                #              'Filing Date', 'Amount', 'Page Number']

                print(('appending {} to worksheet'.format(judg)))
                new_row = [''] * 10
                new_row[0] = judg
                new_row[1] = cj_judgments.get(judg).get('matched_names')
                cj_dob = cj_judgments.get(judg).get('dob')
                if not cj_dob:
                    cj_dob = '-'
                new_row[2] = cj_dob
                new_row[3] = cj_judgments.get(judg).get('address')
                new_row[4] = cj_judgments.get(judg).get('filing_location')[:3]
                new_row[6] = cj_judgments.get(judg).get('filing_date')
                new_row[8] = cj_judgments.get(judg).get('case_amount')
                new_row[9] = cj_judgments.get(judg).get('page_number')
                new_row[cj_idx - 1] = 'X'
                ws.append(new_row)
        return wb

    @staticmethod
    def add_cj_results_to_spreadsheet(csv_filename, cj_judgments, judg_type):
        """

        :param csv_filename:
        :param cj_judgments:
        :param judg_type:
        :return:
        """
        import os
        from django.conf import settings
        from openpyxl import load_workbook, Workbook
        from openpyxl.worksheet import worksheet

        xlsx_filename = csv_filename.replace('.csv', '.xlsx')
        xlsx_pathname = os.path.join(settings.MEDIA_ROOT, xlsx_filename)
        wb = load_workbook(xlsx_pathname)
        ws = wb.worksheets[0]

        judgment_idx = 0
        j_idx = 8
        ws.insert_cols(j_idx)  # add new column for CJ results
        for ws_row in ws.rows: # type: worksheet
            jud_cell_value = ws_row[judgment_idx].value
            if cj_judgments.get(jud_cell_value):
                ws_row[j_idx - 1].value = 'X'
                print(("Found {} match for judgment {}".format(judg_type, jud_cell_value)))
            else:
                print(("No {} match found for judgment {}".format(judg_type, jud_cell_value)))
        ws.cell(row=1, column=j_idx, value=judg_type)  # label header

        wb = JudgmentsExtractor._add_extra_judgments_to_workbook(wb, cj_judgments, judgment_idx, j_idx)

        comp_xlsx_pathname = xlsx_pathname.replace('.xlsx', '-{}.xlsx'.format(judg_type))
        wb.save(comp_xlsx_pathname)
        wb.close()

        # @TODO: need to append CJ judgments at end of result set
        print(("Wrote file to {}".format(comp_xlsx_pathname)))
        return comp_xlsx_pathname

    @staticmethod
    def cj_pdf_comparison(request):
        """
        Merge generated excel spreadsheet data with CJ results by adding a column with new data for comparison
        :param request:
        :return: the completed excel file
        """
        import logging
        import os
        from django.http import FileResponse, Http404
        from orders.models import Order

        order_id = request.POST['order_id']
        # csv_filename = JudgmentsExtractor.get_csv_filename_from_order_id(order_id)

        client_number = '{}-{}'.format(Order.objects.get(id=order_id).title_number, order_id)
        csv_filename = "{}.csv".format(client_number)

        judg_type = None

        if 'cj_file' in request.FILES:
            # if request.FILES['cj_file']:
            judg_type = "CJ"
            cj_saved_pdf = JudgmentsExtractor.handle_uploaded_file(request.FILES['cj_file'])
            # pull CJ results and add to excel file here
            # cj_parser = JudgmentsExtractor(cj_saved_pdf, ptype=judg_type)
            # cj_judgments = cj_parser.parse_pdf_judgments()
            cj_parser = CJPDFExtractor(cj_saved_pdf, ptype="CJ")
            cj_parser.parse_pdf_data_for_excel()
            cj_judgments = cj_parser.parsed_judgments
            comp_xlsx_pathname = JudgmentsExtractor.add_cj_results_to_spreadsheet(csv_filename, cj_judgments, judg_type=judg_type)
        elif 'sc_file' in request.FILES:
            judg_type = "SC"
            sc_saved_pdf = JudgmentsExtractor.handle_uploaded_file(request.FILES['sc_file'])
            # pull SC results and add to excel file here
            sc_parser = JudgmentsExtractor(sc_saved_pdf, ptype=judg_type)
            sc_judgments = sc_parser.parse_pdf_judgments()
            sc_judgments = JudgmentsExtractor.sc_judgments_to_jsn_judgments(sc_judgments)
            comp_xlsx_pathname = JudgmentsExtractor.add_cj_results_to_spreadsheet(csv_filename, sc_judgments, judg_type)

            # logger = logging.getLogger(__name__)
            # logger.info('Sending excel file to browser for download')
            # logger.info('Excel file path is {}'.format(cj_comp_xlsx_pathname))
        try:
            assert(os.path.isfile(comp_xlsx_pathname))
            response = FileResponse(open(comp_xlsx_pathname, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={title}'.format(title='{}-comparison.xlsx'.format(judg_type))
            return response
        except: #  FileNotFoundError:
            raise Http404()

    @staticmethod
    def get_csv_filename_from_order_id(order_id):
        import os
        from django.conf import settings
        from orders.models import Order

        csv_filename = None
        csv_filename_base = Order.objects.get(id=order_id).title_number
        csv_filename = os.path.join(settings.MEDIA_ROOT, csv_filename_base) + '.csv'

        if not os.path.isfile(csv_filename):
            csv_filename = csv_filename.replace('.csv', '.xlsx')

        return csv_filename

    @staticmethod
    def handle_uploaded_file(f):
        import os
        from django.core.files.storage import FileSystemStorage

        pdf_dir = os.path.join('pdftools', 'processed_docs')
        fs = FileSystemStorage(location=pdf_dir)
        filename = fs.save(f.name, f)
        return fs.path(f.name)

    @staticmethod
    def sc_judgment_to_jsn_judgment(j_num):
        """
        Convert a single judgment in state capital format to JSN format
        """
        judg_partial_year = j_num[-2:]
        if int(judg_partial_year) > 90:
            complete_year = '19' + judg_partial_year
        else:
            complete_year = '20' + judg_partial_year
        case_num = j_num.split('-')[0]
        jsn_judgment = case_num[:2] + '-' + case_num[2:] + '-' + complete_year
        jsn_judgment = jsn_judgment.replace(' ', '')
        jsn_judgment = jsn_judgment.replace('*', '')
        return jsn_judgment

    @staticmethod
    def sc_judgments_to_jsn_judgments(d):
        """
        Takes a dictionary of judgments in state capital format and converts them to JSN format
        :param d: dictionary of judgments
        :return: dict dictionary of JSN formatted judgments
        """
        jsn_formatted_dict = {}
        for key, value in list(d.items()):
            new_key = JudgmentsExtractor.sc_judgment_to_jsn_judgment(key)
            jsn_formatted_dict[new_key] = value
        return jsn_formatted_dict

    @staticmethod
    def save_sc_dobs(sc_dob_dict_list, file_hash):
        """
        Parses a list of SC dicts into individual models with a file has
        :param sc_dob_dict_list:
        :return:
        """
        from pdftools.models import SCDobDocument
        from pdftools.utils import JudgmentsExtractor
        from django.core.exceptions import ValidationError

        for d in sc_dob_dict_list:
            parsed_judgment_num = JudgmentsExtractor.sc_judgment_to_jsn_judgment(d['judgment'])
            scdob = SCDobDocument(file_hash=file_hash,
                                  judgment_num=parsed_judgment_num,
                                  party_name=d['name'],
                                  party_dob=d.get('dob', None),
                                  party_ssn=d.get('ssn', None),
                                  party_dlicense=d.get('dlicense', None))
            try:
                scdob.save()
            except ValidationError as e:
                print(('Judgment already exists from a previous document: {}'.format(d)))
                scdob.delete()
                scdob.save()
                print(('Updated judgment {} with new judgment info'.format(parsed_judgment_num)))
            except Exception as e:
                import logging
                print(('Exception while saving data for judgment {}. Continuing...'.format(d['judgment'])))
                logger = logging.getLogger('dobscraper')
                # hdlr = logging.FileHandler('dobscraper.log')
                # logger.addHandler(hdlr)
                logger.setLevel(logging.INFO)
                logger.error('Failed to save entry to database: ' + str(e))


class DobExtractor:

    def __init__(self, pdf_name, ptype):
        """
        Initialize the PDF parser
        :param pdf_name: name of the pdf to parse
        :param ptype: type of parser. CJ for charles jones, JSN or SC for state capital. JSN = CJ
        """
        self.pdf_file = pdf_name
        self.parser_type = ptype


    def parse_pdf_dobs(self):

        pdf_page_offset = 2
        doc_final_judgment_list = []
        with open(self.pdf_file, mode='rb') as f:
            reader = PyPDF2.PdfFileReader(f)
            for i, page in enumerate(reader.pages):
                if i >= pdf_page_offset:
                    page_text = page.extractText()
                    judgment_page_dict_list = self._find_judgments_on_page(page_text, i+pdf_page_offset)
                    if len(judgment_page_dict_list) > 0:
                        # final_judgment_dict.update(x for x in judgment_page_dict_list)
                        doc_final_judgment_list.extend(judgment_page_dict_list)
        print('Parsing complete')
        return doc_final_judgment_list


    def _find_judgments_on_page(self, page_text, pdf_page):
        """
        Parse text for judgments and return matches as dictionary
        :param page_text: extracted text str from pdf
        :param pdf_page: page number of the page being parsed
        :return: list of dicts with matches of judgment data with judgment numbers as keys
        """
        import re
        jdict = None
        jdlist = []
        judgment_type = self.parser_type

        regex = r"JUDGMENT:[\s\S]+?ABSTRACT+?"
        matches = re.findall(regex, page_text, re.MULTILINE)
        # matches = re.finditer(regex, page_text, re.MULTILINE)
        for m in matches:
            jdict = self._parse_party_and_dob(m)
            if self._is_judgment_with_extra_info(jdict):
                jdlist.append(jdict)
            # else:
            #     print("Judgment {} didn't have any extra identifying information. Skipping...".format(jdict.get('judgment')))
        # print 'End of page reached'
        return jdlist

    def _is_judgment_with_extra_info(self, jdict):
        is_valid_judgment = False
        if jdict.get('judgment', None) and jdict.get('name', None) and (jdict.get('dob') or jdict.get('ssn') or jdict.get('dlicense')):
            is_valid_judgment = True
        return is_valid_judgment

    def _parse_party_and_dob(self, match):
        import re
        d = {}
        match_lines = match.split('\n')
        for i, line in enumerate(match_lines):
            if 'JUDGMENT:' in line:
                # original test document
                # j = match_lines[i+1]
                # d['judgment'] = j
                # regex = r'JUDGMENT:([\s\S]+\s+\d+\-\d+)?'
                # regex = r'JUDGMENT:([*]?[\s\S]{2}[\s]*\d{6}\-\d{2})'
                regex = r'JUDGMENT:([*]?[\s\S]{2}[\s]*\d+\-\d+)'
                matches = re.findall(regex, line)
                if matches and len(matches) > 0:
                    try:
                        d['judgment'] = matches[0]
                    except IndexError as e:
                        j = match_lines[i+1]
                        d['judgment'] = j

            if 'DEBTOR(S):' in line:
                # pd = match_lines[i+1]
                # d['debtordob'] = pd
                # regex = r'(DEBTOR[\s\S]+DOB:.+)(\d+\/\d+\/\d+)+?'
                regex = r'(DEBTOR[\s\S]+)DOB:.+?(\d{2}\/\d{2}\/\d{4})+?'
                matches = re.findall(regex, line)
                if matches and len(matches) > 0:
                    debtor_line = matches[0][0]
                    debtor_name = debtor_line.split(':')[1]
                    debtor_name = debtor_name.split(',')[0]  # get rid of commas at the end
                    debtor_dob = matches[0][1]
                    # print('Found debtor and DOB: {} - {}'.format(debtor_name, debtor_dob))
                    d['name'] = debtor_name
                    d['dob'] = debtor_dob

            if 'XXX-XX' in line:
                regex = r'XXX-XX-\d{4}'
                matches = re.findall(regex, line)
                if matches and len(matches) > 0:
                    partial_ssn = matches[0]
                    d['ssn'] = partial_ssn

            if 'LICENSE' in line:
                regex = r'LICENSE\s+=\s+(.{15})'
                matches = re.findall(regex, line)
                if matches and len(matches) > 0:
                    drivers_license = matches[0]
                    d['dlicense'] = drivers_license
        return d


class JudgmentsFileHashType(Enum):
    SCDOB = 'SCDOB'
    ALIAS = 'ALIAS'


class CJPDFExtractor:

    DRIVERS_TOKEN = 'DRIVERS'
    DOB_TOKEN = 'DOB:'
    PDF_PAGE_OFFSET = 1

    class CJExcelData:

        def __init__(self, judgment_number):
            """
            Initialize the data structure for this class
            """
            self.document_number = judgment_number

        document_number = None  # the judgment number
        matched_names = None
        dob = None
        address = None
        filing_location = None
        filing_date = None
        case_amount = None
        page_number = None

    def __init__(self, pdf_name, ptype='CJ'):
        """
        Initialize the PDF parser
        :param pdf_name: name of the pdf to parse
        :param ptype: type of parser. CJ for charles jones, JSN or SC for state capital. JSN = CJ
        """
        self.pdf_file = pdf_name
        self.parser_type = ptype
        self.parsed_judgments = {}  # a dict of CJExcelData judgment data

    def parse_pdf_data_for_excel(self):
        """
        Use the pdf file set during class initialization to extract CJ judgment details
        :return: None
        """
        pdf_page_offset = self.PDF_PAGE_OFFSET
        with open(self.pdf_file, mode='rb') as f:
            reader = PyPDF2.PdfFileReader(f)
            for i, page in enumerate(reader.pages):
                # if i > 2:  # for testing with large pdfs
                #     break
                if i >= pdf_page_offset:
                    page_text = page.extractText()
                    doc_page_number = i + pdf_page_offset
                    page_extracted_judgments = self._extract_judgment_data_from_page(page_text, doc_page_number)
                    self.parsed_judgments.update(page_extracted_judgments)

    def _extract_judgment_data_from_page(self, page_text, doc_page_number):
        """
        Extract judgment details from a page of text extracted from a larger CJ judgment PDF
        :param page_text: page text extracted from PDF
        :param doc_page_number: the page number to use for assignment to the extracted data
        :return: a dict of data extracted from the page text
        """
        excel_data_dict = {}
        abstract_regex = 'SUPERIOR COURT OF NEW JERSEY[\s\S.]*?End of Abstract'
        abstract_matches = re.findall(abstract_regex, page_text)
        for i, abstract_match in enumerate(abstract_matches):
            d = {}
            abstract_text = abstract_match.strip()
            d['document_number'] = self._extract_document_number(abstract_text)
            d['matched_names'] = self._extract_debtor_names(abstract_text)
            d['dob'] = self._extract_dob(abstract_text)
            d['address'] = self._extract_address_section(abstract_text)
            d['filing_location'] = self._extract_filing_location(abstract_text)
            d['filing_date'] = self._extract_filing_date(abstract_text)
            d['case_amount'] = self._extract_case_amount(abstract_text)
            d['page_number'] = doc_page_number - self.PDF_PAGE_OFFSET
            print(d)
            excel_data_dict[d['document_number']] = d

        return excel_data_dict

    def _extract_document_number(self, abstract_match):
        """
        extract the document number from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: judgment/document number
        """
        match_result = None
        regex = r'JUDGMENT NUMBER: (\w+\-\d+\-\d+)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0]
        return match_result

    def _extract_filing_date(self, abstract_match):
        """
        extract filing date from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: filing date
        """
        match_result = None
        regex = r'(DATE DOCKETED|DATE ENTERED): (\d+/\d+/\d+)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0][1]
        return match_result

    def _extract_filing_location(self, abstract_match):
        """
        extract the filing location from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: filing location/venue
        """
        match_result = None
        regex = r'VENUE: (.*)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0]
        return match_result

    def _extract_case_amount(self, abstract_match):
        """
        extract the debt amount from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: the debt amount, stripped of commas
        """
        match_result = None
        regex = r'DEBT:\s\$\s+([\d,\.]+)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0]
            match_result = match_result.replace(',','')
        return match_result

    def _find_name_section(self, abstract_match):
        """
        find the portion of the document containing names
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: section of the document containing debtor name
        """
        match_result = None
        regex = r'DEBTOR\(S\):[\s\S]{,2}\s+(.*)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0]
        return match_result

    def _extract_address_section(self, abstract_match):
        """
        extract the address from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: the address for the abstracted judgment
        """
        match_result = '(No Address)'
        regex = r'.*,.*\d{5}\-\d{,4}'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            match_result = match[0].strip()
        return match_result

    def _extract_aliases(self, abstract_match):
        """
        extract any aliases found in the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: a list of matched aliases
        :rtype: list[str]
        """
        match_result = None
        match_list = []
        regex = r'(A/K/A)[\s\S]{,2}\s+(.+)'
        match = re.findall(regex, abstract_match)
        if len(match) > 0:
            for alias_match in match:
                match_list.append(alias_match[1])
            # match_result = ','.join(match_list)
            match_result = match_list
        return match_result

    def _extract_debtor_names(self, abstract_match):
        """
        extract the debtor names from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: a comma separated list of debtor names and aliases
        """
        match_result = self._find_name_section(abstract_match)
        aliases_result = self._extract_aliases(abstract_match)
        # cleanup match
        # drivers_token = 'DRIVERS'
        # dob_token = 'DOB:'
        if self.DRIVERS_TOKEN in match_result:
            match_result = match_result.split(self.DRIVERS_TOKEN)[0]
        if self.DOB_TOKEN in match_result:
            match_result = match_result.split(self.DOB_TOKEN)[0]
        match_result = match_result.strip()
        if match_result[-1] == ',':
            match_result = match_result[0:-1]
        match_result = match_result.strip()
        if aliases_result:
            # match_result = ','.join([aliases_result])
            match_result_list = [match_result] + aliases_result
            match_result = ','.join(set(match_result_list))
        return match_result

    def _extract_dob(self, abstract_match):
        """
        extract the date of birth from the supplied data
        :param abstract_match: text for a single abstract extracted from the larger CJ pdf
        :return: the date of birth
        """
        match_result = self._find_name_section(abstract_match)
        if self.DOB_TOKEN in match_result:
            match_result = match_result.split(self.DOB_TOKEN)[1]
        else:
            match_result = '-'
        return match_result

